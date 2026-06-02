import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

# Load the original workbook to get the default data
wb_orig = openpyxl.load_workbook('American Method Latest 28052026.xlsx', data_only=True)
ws_orig = wb_orig['5 MAJN']

stns = []
exg = []
pro = []
for r in range(3, 100):
    stns.append(ws_orig.cell(row=r, column=1).value)
    exg.append(ws_orig.cell(row=r, column=2).value or 0)
    pro.append(ws_orig.cell(row=r, column=3).value or 0)

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Curve Realignment'
ws.views.sheetView[0].showGridLines = True

# Colors
navy_blue = '1F4E79'
light_gray = 'F2F2F2'
light_blue = 'DDEBF7'
white = 'FFFFFF'

# Fonts
font_title = Font(name='Segoe UI', size=16, bold=True, color='1F4E79')
font_subtitle = Font(name='Segoe UI', size=10, italic=True, color='595959')
font_header = Font(name='Segoe UI', size=11, bold=True, color=white)
font_data = Font(name='Segoe UI', size=10)
font_total = Font(name='Segoe UI', size=11, bold=True)
font_slew = Font(name='Segoe UI', size=10, bold=True, color='1F4E79')

# Fills
fill_header = PatternFill(start_color=navy_blue, end_color=navy_blue, fill_type='solid')
fill_zebra = PatternFill(start_color=light_gray, end_color=light_gray, fill_type='solid')
fill_slew = PatternFill(start_color=light_blue, end_color=light_blue, fill_type='solid')
fill_total = PatternFill(start_color='E6EEF8', end_color='E6EEF8', fill_type='solid')

# Borders
thin_side = Side(border_style='thin', color='D9D9D9')
border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_header = Border(left=thin_side, right=thin_side, top=Side(border_style='medium', color=navy_blue), bottom=Side(border_style='medium', color=navy_blue))
border_total = Border(top=Side(border_style='thin', color='000000'), bottom=Side(border_style='double', color='000000'))

# Alignments
align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_right = Alignment(horizontal='right', vertical='center')

# Title block
ws.merge_cells('A1:I1')
ws['A1'] = 'TRACK CURVE REALIGNMENT SHEET (HALLADE METHOD)'
ws['A1'].font = font_title
ws['A1'].alignment = align_center

ws.merge_cells('A2:I2')
ws['A2'] = 'Dynamic Slew Calculator | Paste your curve data below (supports up to 500 stations)'
ws['A2'].font = font_subtitle
ws['A2'].alignment = align_center

# Column Headers
headers = [
    'Stn No', "Existing Versine (B)\n(mm)", "Proposed Versine (C)\n(mm)", 
    "Versine Diff (D)\n(=C-B)", "First Sum (E)\n(=E[prev]+D)", "Second Sum (F)\n(=F[prev]+E)", 
    "Raw Slew (G)\n(=-2*F)", "Linear Correction (H)\n(=G[end]*dist/length)", "Corrected Slew (I)\n(=G-H)"
]

for col_idx, text in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx)
    cell.value = text
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_header
ws.row_dimensions[4].height = 40

start_row = 5
max_stations = 500
end_row = start_row + max_stations - 1

# Populate up to max_stations rows with dynamic formulas
for idx in range(max_stations):
    r = start_row + idx
    
    # Check if we have default data for this row
    has_data = idx < len(stns)
    stn_val = stns[idx] if has_data else ""
    exg_val = exg[idx] if has_data else ""
    pro_val = pro[idx] if has_data else ""
    
    # Values
    ws.cell(row=r, column=1, value=stn_val).alignment = align_center
    if has_data:
        ws.cell(row=r, column=2, value=exg_val).number_format = '#,##0'
        ws.cell(row=r, column=3, value=pro_val).number_format = '#,##0'
    else:
        ws.cell(row=r, column=2, value="").number_format = '#,##0'
        ws.cell(row=r, column=3, value="").number_format = '#,##0'
    
    # Dynamic Formulas wrapped in IF statement
    ws.cell(row=r, column=4, value=f'=IF(A{r}="","",C{r}-B{r})').number_format = '#,##0'
    
    if r == start_row:
        ws.cell(row=r, column=5, value=f'=IF(A{r}="","",D{r})').number_format = '#,##0'
        ws.cell(row=r, column=6, value=f'=IF(A{r}="","",E{r})').number_format = '#,##0'
    else:
        ws.cell(row=r, column=5, value=f'=IF(A{r}="","",E{r-1}+D{r})').number_format = '#,##0'
        ws.cell(row=r, column=6, value=f'=IF(A{r}="","",F{r-1}+E{r})').number_format = '#,##0'
        
    ws.cell(row=r, column=7, value=f'=IF(A{r}="","",-2*F{r})').number_format = '#,##0.0'
    ws.cell(row=r, column=8, value=f'=IF(A{r}="","",INDEX($G$5:$G${end_row},COUNTA($A$5:$A${end_row}))*(ROW()-{start_row})/(COUNTA($A$5:$A${end_row})-1))').number_format = '#,##0.0'
    ws.cell(row=r, column=9, value=f'=IF(A{r}="","",G{r}-H{r})').number_format = '#,##0.0'
    
    # Formatting
    for c in range(1, 10):
        cell = ws.cell(row=r, column=c)
        cell.font = font_data
        cell.border = border_cell
        if c > 1:
            cell.alignment = align_right
        
        # Zebra striping and highlight slew
        if c == 9:
            cell.fill = fill_slew
            cell.font = font_slew
        elif r % 2 == 0:
            cell.fill = fill_zebra
            
    ws.row_dimensions[r].height = 20

# Totals Row
tot_row = end_row + 2
ws.cell(row=tot_row, column=1, value='Total').font = font_total
ws.cell(row=tot_row, column=1).alignment = align_center

ws.cell(row=tot_row, column=2, value=f'=SUM(B{start_row}:B{end_row})').number_format = '#,##0'
ws.cell(row=tot_row, column=3, value=f'=SUM(C{start_row}:C{end_row})').number_format = '#,##0'
ws.cell(row=tot_row, column=4, value=f'=SUM(D{start_row}:D{end_row})').number_format = '#,##0'

for c in range(1, 10):
    cell = ws.cell(row=tot_row, column=c)
    cell.font = font_total
    cell.border = border_total
    cell.fill = fill_total
    if c > 1 and cell.value:
        cell.alignment = align_right
ws.row_dimensions[tot_row].height = 24

# Freeze panes
ws.freeze_panes = 'A5'

# Auto-adjust column widths
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.row < 4:
            continue
        val_str = str(cell.value or '')
        if '\n' in val_str:
            lines = val_str.split('\n')
            max_len = max(max_len, max(len(l) for l in lines))
        else:
            max_len = max(max_len, len(val_str))
    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# --- Charts Worksheet ---
ws_plots = wb.create_sheet(title='Plots')
ws_plots.views.sheetView[0].showGridLines = True

# Title for Plots Sheet
ws_plots.merge_cells('A1:L1')
ws_plots['A1'] = 'TRACK GEOMETRY VISUALIZATION'
ws_plots['A1'].font = Font(name='Segoe UI', size=16, bold=True, color='1F4E79')
ws_plots['A1'].alignment = align_center
ws_plots.row_dimensions[1].height = 30

active_end_row = start_row + len(stns) - 1

# Chart 1: Versines
chart_ver = LineChart()
chart_ver.title = 'Versine Profile (Existing vs. Proposed)'
chart_ver.style = 13
chart_ver.y_axis.title = 'Versine (mm)'
chart_ver.x_axis.title = 'Station No'
chart_ver.width = 24
chart_ver.height = 14

data_ver = Reference(ws, min_col=2, min_row=4, max_col=3, max_row=active_end_row)
cats_ver = Reference(ws, min_col=1, min_row=5, max_row=active_end_row)
chart_ver.add_data(data_ver, titles_from_data=True)
chart_ver.set_categories(cats_ver)

# Chart 1 Styling
chart_ver.series[0].graphicalProperties.line.solidFill = 'FF5B5B' # Red for existing
chart_ver.series[0].graphicalProperties.line.width = 25000
chart_ver.series[1].graphicalProperties.line.solidFill = '5B9BD5' # Blue for proposed
chart_ver.series[1].graphicalProperties.line.width = 30000

# Chart 2: Slew
chart_slew = LineChart()
chart_slew.title = 'Slew Profile (Required Shift)'
chart_slew.style = 13
chart_slew.y_axis.title = 'Slew (mm)'
chart_slew.x_axis.title = 'Station No'
chart_slew.width = 24
chart_slew.height = 14

data_slew = Reference(ws, min_col=9, min_row=4, max_col=9, max_row=active_end_row)
cats_slew = cats_ver
chart_slew.add_data(data_slew, titles_from_data=True)
chart_slew.set_categories(cats_slew)

# Chart 2 Styling
chart_slew.series[0].graphicalProperties.line.solidFill = '2E75B6' # Dark Blue for slew
chart_slew.series[0].graphicalProperties.line.width = 30000

# Place charts in Plots sheet
ws_plots.add_chart(chart_ver, 'A3')
ws_plots.add_chart(chart_slew, 'A31')

# Save
wb.save('Hallade_Curve_Realignment.xlsx')
print('Workbook with plots saved successfully')
