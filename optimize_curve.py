import openpyxl
import numpy as np
from scipy.optimize import minimize
import os

def run_optimization():
    print("="*60)
    print("        RAILWAY CURVE REALIGNMENT OPTIMIZATION TOOL")
    print("        Method: Hallade Double-Summation (Exact)")
    print("        Supports: Simple, Compound, and Reverse Curves")
    print("="*60)
    
    file_name = 'Hallade_Curve_Realignment.xlsx'
    if not os.path.exists(file_name):
        print(f"Error: Could not find '{file_name}' in the current directory.")
        return
        
    print(f"Loading '{file_name}'...")
    wb = openpyxl.load_workbook(file_name, data_only=True)
    ws = wb['Curve Realignment']
    
    # Read active data from the sheet
    stns = []
    exg = []
    start_row = 5
    for r in range(start_row, 505):
        stn = ws.cell(row=r, column=1).value
        if stn is None or str(stn).strip() == "":
            break
        exg_val = ws.cell(row=r, column=2).value or 0
        stns.append(stn)
        exg.append(exg_val)
        
    N = len(stns)
    if N < 5:
        print(f"Error: Found only {N} stations. Need at least 5 stations to optimize.")
        return
        
    print(f"Detected {N} active stations (Row 5 to {start_row + N - 1}).")
    v_ex = np.array(exg, dtype=float)
    
    # Identify active stations (non-zero curve region)
    # We automatically detect the first and last non-zero existing versine
    non_zero_indices = np.where(np.abs(v_ex) > 0)[0]
    if len(non_zero_indices) == 0:
        print("Warning: All existing versines are 0. Nothing to optimize.")
        return
        
    first_active = max(0, non_zero_indices[0] - 3)
    last_active = min(N - 1, non_zero_indices[-1] + 3)
    
    active_mask = np.zeros(N, dtype=bool)
    active_mask[first_active:last_active+1] = True
    active_indices = np.where(active_mask)[0]
    n_active = len(active_indices)
    
    print(f"Curve zone for optimization: Stations {stns[first_active]} to {stns[last_active]} ({n_active} active stations).")

    # Ask the user for a slew limit
    print("-" * 60)
    user_input = input("Enter desired maximum slew limit in mm\n(or press Enter to automatically find the minimum feasible limit): ").strip()
    
    target_limit = None
    if user_input != "":
        try:
            target_limit = float(user_input)
            if target_limit <= 0:
                print("Error: Slew limit must be a positive number.")
                return
        except ValueError:
            print("Error: Invalid number entered.")
            return

    # Optimization functions
    def objective(x_active):
        x = np.zeros(N)
        x[active_indices] = x_active
        return np.sum(np.diff(x, n=2)**2)

    def constr_sum(x_active):
        x = np.zeros(N)
        x[active_indices] = x_active
        return np.sum(x) - np.sum(v_ex)

    def constr_moment(x_active):
        x = np.zeros(N)
        x[active_indices] = x_active
        d = x - v_ex
        weights = np.arange(N, 0, -1)
        return np.sum(weights * d)

    def get_slew(x_active):
        x = np.zeros(N)
        x[active_indices] = x_active
        d = x - v_ex
        s = np.cumsum(d)
        m = np.cumsum(s)
        return -2 * m

    # Dynamically set bounds based on existing versine direction (positive or negative)
    bounds = []
    for idx in active_indices:
        if v_ex[idx] >= 0:
            bounds.append((0, 300))
        else:
            bounds.append((-300, 0))

    x0_active = v_ex[active_indices]

    def solve_for_limit(limit):
        constraints = [
            {'type': 'eq', 'fun': constr_sum},
            {'type': 'eq', 'fun': constr_moment}
        ]
        for i in range(N):
            constraints.append({'type': 'ineq', 'fun': (lambda idx=i, lim=limit: lambda x_act: lim - get_slew(x_act)[idx])()})
            constraints.append({'type': 'ineq', 'fun': (lambda idx=i, lim=limit: lambda x_act: get_slew(x_act)[idx] + lim)()})
            
        return minimize(objective, x0_active, method='SLSQP', bounds=bounds, constraints=constraints, options={'maxiter': 500, 'ftol': 1e-6})

    res = None
    final_limit = None
    
    if target_limit is not None:
        print(f"Solving optimization with maximum slew limit = {target_limit:.1f} mm...")
        res = solve_for_limit(target_limit)
        if res.success:
            final_limit = target_limit
        else:
            print(f"Warning: Optimization failed with strict {target_limit} mm limit.")
            print("Searching for a feasible limit instead...")
            target_limit = None

    if target_limit is None:
        # Automatic search for smallest feasible limit
        for limit in [10.0, 15.0, 20.0, 25.0, 30.0, 35.0, 40.0, 45.0, 50.0, 60.0, 75.0, 100.0, 150.0, 200.0]:
            print(f"Testing feasibility with slew limit = {limit:.1f} mm...")
            res_test = solve_for_limit(limit)
            if res_test.success:
                print(f"  Feasible solution found at {limit:.1f} mm limit!")
                res = res_test
                final_limit = limit
                break
            else:
                print(f"  Not feasible at {limit:.1f} mm.")

    if res is None or not res.success:
        print("\nError: Could not find a feasible solution. Check data balance.")
        return

    # Extract continuous solution
    x_opt = np.zeros(N)
    x_opt[active_indices] = res.x

    # Integer refinement (coordinate descent) to eliminate rounding error accumulation
    print("Refining solution to whole millimeters (integers)...")
    current_x = np.round(x_opt).astype(int)
    
    # Ensure total sum is exactly balanced
    diff_sum = int(np.sum(current_x) - np.sum(v_ex))
    if diff_sum != 0:
        step = 1 if diff_sum < 0 else -1
        for _ in range(abs(diff_sum)):
            # Distribute the discrepancy to the index with largest absolute value
            best_idx = active_indices[np.argmax(np.abs(current_x[active_indices]))]
            current_x[best_idx] += step

    def calc_max_slew(x):
        d = x - v_ex
        s = np.cumsum(d)
        m = np.cumsum(s)
        raw_slew = -2 * m
        corr = raw_slew[-1] * np.arange(N) / (N - 1)
        slew = raw_slew - corr
        return np.max(np.abs(slew))

    def calc_smoothness(x):
        return np.sum(np.diff(x, n=2)**2)

    # Local search under reverse curve bounds
    current_cost = calc_max_slew(current_x) + 0.005 * calc_smoothness(current_x)
    improved = True
    iterations = 0
    while improved and iterations < 1000:
        improved = False
        for idx1 in active_indices:
            for idx2 in active_indices:
                if idx1 == idx2:
                    continue
                test_x = current_x.copy()
                test_x[idx1] += 1
                test_x[idx2] -= 1
                
                # Check dynamic direction bounds
                lim1 = (0, 300) if v_ex[idx1] >= 0 else (-300, 0)
                lim2 = (0, 300) if v_ex[idx2] >= 0 else (-300, 0)
                if not (lim1[0] <= test_x[idx1] <= lim1[1] and lim2[0] <= test_x[idx2] <= lim2[1]):
                    continue
                    
                test_cost = calc_max_slew(test_x) + 0.005 * calc_smoothness(test_x)
                if test_cost < current_cost - 0.001:
                    current_x = test_x
                    current_cost = test_cost
                    improved = True
                    break
            if improved:
                break
        iterations += 1

    final_max_slew = calc_max_slew(current_x)
    print(f"\nOptimization Results:")
    print(f"  Maximum Slew: {final_max_slew:.1f} mm")
    print(f"  Proposed Curve Smoothness Cost: {calc_smoothness(current_x):.0f}")
    
    # Save back to Excel
    print(f"Saving optimized proposed versines back to '{file_name}'...")
    wb_write = openpyxl.load_workbook(file_name)
    ws_write = wb_write['Curve Realignment']
    for idx in range(N):
        r = start_row + idx
        ws_write.cell(row=r, column=3, value=int(current_x[idx]))
        
    wb_write.save(file_name)
    print("Workbook saved successfully!")
    print("="*60)
    print("You can now return to Excel. The charts and slews have updated automatically.")
    print("="*60)

if __name__ == '__main__':
    run_optimization()
